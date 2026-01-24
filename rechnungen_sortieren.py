import argparse
import os
import re
import subprocess
import sys
import traceback
import zlib
from decimal import Decimal, ROUND_HALF_EVEN, getcontext

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

getcontext().prec = 12

MONEY = Decimal("0.01")
TOLERANZ = Decimal("0.01")


def quantize_money(value):
    return value.quantize(MONEY, rounding=ROUND_HALF_EVEN)


def normalize(text):
    return " ".join(str(text).lower().strip().split())


def parse_decimal(value):
    return Decimal(str(value).replace(",", "."))


def extract_positions_sequential(text, artikel_namen):
    rest = normalize(text)
    gefunden = []

    while rest:
        match = None
        for name in artikel_namen:
            if rest.startswith(name):
                match = name
                break

        if not match:
            break

        gefunden.append(match)
        rest = rest[len(match):].lstrip()
        if rest.startswith(","):
            rest = rest[1:].lstrip()

    return gefunden, rest


def strip_quantities(text):
    return re.sub(r"\s*\(\d+x\)", "", text)


def parse_pdf_totals(pdf_path):
    pdf = open(pdf_path, "rb").read()
    objects = {}
    for match in re.finditer(rb"(\d+)\s+0\s+obj", pdf):
        num = int(match.group(1))
        start = match.start()
        end = pdf.find(b"endobj", start)
        if end == -1:
            continue
        objects[num] = pdf[start : end + 6]

    def parse_cmap(obj_num):
        obj = objects.get(obj_num)
        if not obj or b"stream" not in obj:
            return {}
        m = re.search(rb"stream\r?\n", obj)
        if not m:
            return {}
        start = m.end()
        end = obj.find(b"endstream", start)
        data = obj[start:end]
        try:
            dec = zlib.decompress(data)
        except Exception:
            dec = data
        text = dec.decode("latin-1")
        mapping = {}
        for block in re.findall(r"beginbfchar(.*?)endbfchar", text, re.S):
            for src, dst in re.findall(r"<([0-9A-Fa-f]+)>\s*<([0-9A-Fa-f]+)>", block):
                cid = int(src, 16)
                try:
                    mapping[cid] = chr(int(dst, 16))
                except ValueError:
                    pass
        for block in re.findall(r"beginbfrange(.*?)endbfrange", text, re.S):
            for line in block.strip().splitlines():
                line = line.strip()
                if not line:
                    continue
                m = re.match(
                    r"<([0-9A-Fa-f]+)>\s*<([0-9A-Fa-f]+)>\s*<([0-9A-Fa-f]+)>",
                    line,
                )
                if m:
                    start_c = int(m.group(1), 16)
                    end_c = int(m.group(2), 16)
                    dst = int(m.group(3), 16)
                    for cid in range(start_c, end_c + 1):
                        try:
                            mapping[cid] = chr(dst + (cid - start_c))
                        except ValueError:
                            pass
                    continue
                m = re.match(
                    r"<([0-9A-Fa-f]+)>\s*<([0-9A-Fa-f]+)>\s*\[(.*)\]",
                    line,
                )
                if m:
                    start_c = int(m.group(1), 16)
                    end_c = int(m.group(2), 16)
                    dsts = re.findall(r"<([0-9A-Fa-f]+)>", m.group(3))
                    for i, dst in enumerate(dsts):
                        cid = start_c + i
                        if cid > end_c:
                            break
                        try:
                            mapping[cid] = chr(int(dst, 16))
                        except ValueError:
                            pass
        return mapping

    font_obj_to_cmap = {}
    for num, obj in objects.items():
        if b"/Type /Font" in obj and b"/ToUnicode" in obj:
            m = re.search(rb"/ToUnicode\s+(\d+)\s+0\s+R", obj)
            if m:
                cmap_obj = int(m.group(1))
                font_obj_to_cmap[num] = parse_cmap(cmap_obj)

    font_name_to_obj = {}
    for obj in objects.values():
        if b"/Font" in obj and b"<<" in obj:
            for m in re.finditer(rb"/Font\s*<<([^>]+)>>", obj, re.S):
                block = m.group(1)
                for fm in re.finditer(rb"/(\w+)\s+(\d+)\s+0\s+R", block):
                    name = fm.group(1).decode("latin-1")
                    objnum = int(fm.group(2))
                    font_name_to_obj[name] = objnum

    def decode_hex(hexstr, cmap):
        chars = []
        for i in range(0, len(hexstr), 4):
            cid_hex = hexstr[i : i + 4]
            if len(cid_hex) < 4:
                continue
            cid = int(cid_hex, 16)
            ch = cmap.get(cid)
            if ch:
                chars.append(ch)
        return "".join(chars)

    def tokenize(data):
        i = 0
        n = len(data)
        while i < n:
            c = data[i : i + 1]
            if c in b" \t\r\n\f\x00":
                i += 1
                continue
            if c == b"%":
                j = data.find(b"\n", i)
                if j == -1:
                    break
                i = j + 1
                continue
            if c == b"/":
                j = i + 1
                while j < n and data[j : j + 1] not in b" \t\r\n\f/[]()<>%":
                    j += 1
                yield ("name", data[i:j].decode("latin-1"))
                i = j
                continue
            if c == b"(":
                j = i + 1
                depth = 1
                buf = []
                while j < n and depth > 0:
                    ch = data[j : j + 1]
                    if ch == b"\\":
                        if j + 1 < n:
                            buf.append(data[j + 1 : j + 2])
                            j += 2
                            continue
                    if ch == b"(":
                        depth += 1
                    elif ch == b")":
                        depth -= 1
                        if depth == 0:
                            j += 1
                            break
                    buf.append(ch)
                    j += 1
                yield ("string", bytes(buf))
                i = j
                continue
            if c == b"<":
                if i + 1 < n and data[i + 1 : i + 2] == b"<":
                    yield ("dict_start", "<<")
                    i += 2
                    continue
                j = data.find(b">", i)
                if j == -1:
                    break
                hexstr = data[i + 1 : j].decode("latin-1").strip()
                yield ("hex", hexstr)
                i = j + 1
                continue
            if c == b">":
                if i + 1 < n and data[i + 1 : i + 2] == b">":
                    yield ("dict_end", ">>")
                    i += 2
                    continue
                i += 1
                continue
            if c == b"[":
                yield ("array_start", "[")
                i += 1
                continue
            if c == b"]":
                yield ("array_end", "]")
                i += 1
                continue
            if c in b"+-0123456789.":
                j = i + 1
                while j < n and data[j : j + 1] in b"+-0123456789.":
                    j += 1
                token = data[i:j].decode("latin-1")
                try:
                    val = float(token)
                except ValueError:
                    val = 0.0
                yield ("number", val)
                i = j
                continue
            j = i + 1
            while j < n and data[j : j + 1] not in b" \t\r\n\f/[]()<>%":
                j += 1
            yield ("op", data[i:j].decode("latin-1"))
            i = j

    text_items = []
    for obj in objects.values():
        if b"/Length" not in obj or b"stream" not in obj:
            continue
        m = re.search(rb"stream\r?\n", obj)
        if not m:
            continue
        start = m.end()
        end = obj.find(b"endstream", start)
        data = obj[start:end]
        try:
            dec = zlib.decompress(data)
        except Exception:
            continue
        if b"Tj" not in dec and b"TJ" not in dec:
            continue
        text_matrix = [1, 0, 0, 1, 0, 0]
        stack = []
        current_font = None
        tokens = list(tokenize(dec))
        i = 0
        while i < len(tokens):
            typ, val = tokens[i]
            if typ in (
                "number",
                "name",
                "hex",
                "string",
                "array_start",
                "array_end",
                "dict_start",
                "dict_end",
            ):
                stack.append((typ, val))
                i += 1
                continue
            if typ == "op":
                op = val
                if op == "Tf" and len(stack) >= 2:
                    font_name = stack[-2][1]
                    current_font = (
                        font_name.lstrip("/") if isinstance(font_name, str) else font_name
                    )
                    stack = []
                elif op == "Tm" and len(stack) >= 6:
                    nums = [s[1] for s in stack[-6:]]
                    text_matrix = [
                        nums[0],
                        nums[1],
                        nums[2],
                        nums[3],
                        nums[4],
                        nums[5],
                    ]
                    stack = []
                elif op == "Td" and len(stack) >= 2:
                    tx = stack[-2][1]
                    ty = stack[-1][1]
                    text_matrix[4] += tx
                    text_matrix[5] += ty
                    stack = []
                elif op == "Tj" and len(stack) >= 1:
                    item = stack[-1]
                    cmap = {}
                    if current_font in font_name_to_obj:
                        cmap = font_obj_to_cmap.get(font_name_to_obj[current_font], {})
                    text = ""
                    if item[0] == "hex":
                        text = decode_hex(item[1], cmap)
                    elif item[0] == "string":
                        text = item[1].decode("latin-1")
                    if text:
                        x = text_matrix[4]
                        y = text_matrix[5]
                        text_items.append((x, y, text))
                    stack = []
                elif op == "TJ" and len(stack) >= 1:
                    arr = []
                    idx = None
                    for j in range(len(stack) - 1, -1, -1):
                        if stack[j][0] == "array_start":
                            idx = j
                            break
                    if idx is not None:
                        arr = stack[idx + 1 :]
                    parts = []
                    cmap = {}
                    if current_font in font_name_to_obj:
                        cmap = font_obj_to_cmap.get(font_name_to_obj[current_font], {})
                    for typ2, val2 in arr:
                        if typ2 == "hex":
                            parts.append(decode_hex(val2, cmap))
                        elif typ2 == "string":
                            parts.append(val2.decode("latin-1"))
                    text = "".join(parts)
                    if text:
                        x = text_matrix[4]
                        y = text_matrix[5]
                        text_items.append((x, y, text))
                    stack = []
                else:
                    stack = []
                i += 1
                continue
            i += 1

    lines = {}
    for x, y, text in text_items:
        y_key = round(y, 1)
        lines.setdefault(y_key, []).append((x, text))

    sorted_lines = []
    for y in sorted(lines.keys(), reverse=True):
        parts = sorted(lines[y], key=lambda t: t[0])
        line = " ".join([p[1] for p in parts])
        sorted_lines.append(line)

    def cleanup(line):
        marker = "@@"
        line = re.sub(r"\s{2,}", marker, line.strip())
        line = line.replace(" ", "")
        line = line.replace(marker, " ")
        return line

    cleaned = [cleanup(l) for l in sorted_lines]

    invoice_blocks = []
    current = None
    for line in cleaned:
        m = re.match(r"^(\d{5})", line)
        if m:
            if current:
                invoice_blocks.append(current)
            current = {"invoice": m.group(1), "lines": [line]}
        elif current:
            current["lines"].append(line)
    if current:
        invoice_blocks.append(current)

    pdf_data = {}
    for block in invoice_blocks:
        inv = block["invoice"]
        entries = []
        in_table = False
        for line in block["lines"]:
            if line.startswith("USt %Netto"):
                in_table = True
                continue
            if in_table:
                if line.startswith("Transaktion:"):
                    break
                rate_match = re.match(r"^[A-Z]+(19|7)", line)
                rate = None
                remainder = line
                if rate_match:
                    rate = rate_match.group(1)
                    remainder = line[rate_match.end() :]
                amounts = re.findall(r"\d{1,3},\d{2}", remainder)
                if len(amounts) == 3 and rate:
                    entries.append(
                        {
                            "rate": rate,
                            "net": amounts[0],
                            "tax": amounts[1],
                            "gross": amounts[2],
                        }
                    )
        if entries:
            pdf_data[inv] = {
                e["rate"]: {
                    "net": parse_decimal(e["net"]),
                    "tax": parse_decimal(e["tax"]),
                    "gross": parse_decimal(e["gross"]),
                }
                for e in entries
            }

    return pdf_data


def main():
    parser = argparse.ArgumentParser(description="Rechnungen sortieren und auswerten.")
    parser.add_argument("--rechnungen", help="Pfad zur Rechnungs-CSV")
    parser.add_argument("--artikel", help="Pfad zur Artikelliste-CSV")
    parser.add_argument(
        "--sammelrechnung",
        help="Pfad zur Sammelrechnung-PDF (optional, für exakte Summen)",
    )
    parser.add_argument(
        "--output",
        help="Pfad zur Ausgabe-Exceldatei (optional, Standard: <Rechnung>_sortiert.xlsx)",
    )
    parser.add_argument(
        "--no-gui",
        action="store_true",
        help="Kein GUI verwenden (Pfadangaben erforderlich).",
    )
    args = parser.parse_args()

    rechnung_path = args.rechnungen
    artikel_path = args.artikel
    sammelrechnung_path = args.sammelrechnung

    if not rechnung_path or not artikel_path or not sammelrechnung_path:
        if args.no_gui:
            raise SystemExit(
                "Bitte --rechnungen, --artikel und --sammelrechnung angeben (oder GUI nutzen)."
            )
        import tkinter as tk
        from tkinter import filedialog, messagebox

        root = tk.Tk()
        root.withdraw()

        rechnung_path = filedialog.askopenfilename(
            title="Wähle die Rechnungsdatei aus",
            filetypes=[("CSV Dateien", "*.csv")],
        )
        if not rechnung_path:
            return

        artikel_path = filedialog.askopenfilename(
            title="Bitte Artikelliste auswählen",
            filetypes=[("CSV Dateien", "*.csv")],
        )
        if not artikel_path:
            return

        sammelrechnung_path = filedialog.askopenfilename(
            title="Bitte Sammelrechnung (PDF) auswählen",
            filetypes=[("PDF Dateien", "*.pdf")],
        )
        if not sammelrechnung_path:
            return

    try:
        rechnung_df = pd.read_csv(rechnung_path, sep=";", dtype=str)
        artikel_df = pd.read_csv(artikel_path, sep=";", dtype=str)
        pdf_totals = parse_pdf_totals(sammelrechnung_path)

        artikel_df["Name_norm"] = artikel_df["Name*"].apply(normalize)

        artikel_namen = sorted(
            artikel_df["Name_norm"].tolist(),
            key=len,
            reverse=True
        )

        artikel_map = {
            row["Name_norm"]: row
            for _, row in artikel_df.iterrows()
        }

        neue_zeilen = []

        for _, row in rechnung_df.iterrows():
            ust_werte = {
                u.strip()
                for u in str(row["Position(en) USt %"]).split(",")
                if u.strip()
            }

            if len(ust_werte) <= 1:
                neue_zeilen.append((row, None))
                continue

            artikel_seq, _ = extract_positions_sequential(
                row["Position(en)"],
                artikel_namen
            )

            mengen = [int(x.strip()) for x in row["Position(en) Anzahl"].split("|")]

            artikel_7 = []
            artikel_19 = []
            artikel_0 = []

            mengen_7 = []
            mengen_19 = []
            mengen_0 = []

            netto_7 = Decimal("0.00")
            steuer_7_sum = Decimal("0.00")
            brutto_7_sum = Decimal("0.00")
            netto_19 = Decimal("0.00")
            steuer_19_sum = Decimal("0.00")
            brutto_19_sum = Decimal("0.00")
            brutto_0_sum = Decimal("0.00")
            has_spende = False

            for artikel_norm, menge in zip(artikel_seq, mengen):
                daten = artikel_map.get(artikel_norm)
                if daten is None:
                    continue

                ust_num = "".join(c for c in str(daten["USt.*"]) if c.isdigit())
                preis_brutto = quantize_money(
                    parse_decimal(daten["Verkaufspreis Brutto*"]) * Decimal(menge)
                )

                if ust_num == "0":
                    original_name = daten["Name*"]
                    artikel_0.append(f"{original_name} ({menge}x)")
                    mengen_0.append(str(menge))
                    brutto_0_sum += preis_brutto
                    has_spende = True
                    continue
                steuersatz = Decimal(ust_num) / Decimal("100")
                preis_netto = quantize_money(preis_brutto / (Decimal("1") + steuersatz))
                steuer = quantize_money(preis_brutto - preis_netto)
                brutto = preis_brutto
                    
                original_name = daten["Name*"]

                if ust_num == "7":
                    artikel_7.append(f"{original_name} ({menge}x)")
                    mengen_7.append(str(menge))
                    netto_7 += preis_netto
                    steuer_7_sum += steuer
                    brutto_7_sum += brutto

                elif ust_num == "19":
                    artikel_19.append(f"{original_name} ({menge}x)")
                    mengen_19.append(str(menge))
                    netto_19 += preis_netto
                    steuer_19_sum += steuer
                    brutto_19_sum += brutto

            original_brutto = parse_decimal(row["Brutto €"])

            if pdf_totals.get(row["Nr."]):
                target = pdf_totals[row["Nr."]]
                if "7" in target:
                    netto_7 = target["7"]["net"]
                    steuer_7_sum = target["7"]["tax"]
                    brutto_7_sum = target["7"]["gross"]
                if "19" in target:
                    netto_19 = target["19"]["net"]
                    steuer_19_sum = target["19"]["tax"]
                    brutto_19_sum = target["19"]["gross"]

            if has_spende:
                brutto_0_sum = quantize_money(
                    original_brutto - brutto_7_sum - brutto_19_sum
                )

            berechnetes_brutto = quantize_money(brutto_7_sum + brutto_19_sum + brutto_0_sum)

            spendenbetrag = quantize_money(original_brutto - berechnetes_brutto)

            brutto_ok = abs(original_brutto - berechnetes_brutto - spendenbetrag) <= TOLERANZ
            status = "geteilt_ok" if brutto_ok else "geteilt_fehler"

            if artikel_7:
                steuer_7 = quantize_money(steuer_7_sum)
                brutto_7 = quantize_money(brutto_7_sum)

                neu = row.copy()
                neu["Position(en)"] = ", ".join(artikel_7)
                
                if len(mengen_7) == 1:
                    neu["Position(en) Anzahl"] = mengen_7[0]
                else:
                    neu["Position(en) Anzahl"] = "| ".join(mengen_7)
                
                neu["Position(en) USt %"] = "7"
                neu["Netto €"] = f"{netto_7:.2f}".replace(".", ",")
                neu["Steuer €"] = f"{steuer_7:.2f}".replace(".", ",")
                neu["Brutto €"] = f"{brutto_7:.2f}".replace(".", ",")
                neue_zeilen.append((neu, status))

            if artikel_19:
                steuer_19 = quantize_money(steuer_19_sum)
                brutto_19 = quantize_money(brutto_19_sum)

                neu = row.copy()
                neu["Position(en)"] = ", ".join(artikel_19)
                
                if len(mengen_19) == 1:
                    neu["Position(en) Anzahl"] = mengen_19[0]
                else:
                    neu["Position(en) Anzahl"] = "| ".join(mengen_19)
                
                neu["Position(en) USt %"] = "19"
                neu["Netto €"] = f"{netto_19:.2f}".replace(".", ",")
                neu["Steuer €"] = f"{steuer_19:.2f}".replace(".", ",")
                neu["Brutto €"] = f"{brutto_19:.2f}".replace(".", ",")
                neue_zeilen.append((neu, status))

            if not sammelrechnung_path and abs(spendenbetrag) > TOLERANZ:
                neu = row.copy()
                neu["Position(en)"] = "Spende"
                neu["Position(en) USt %"] = "0"
                neu["Netto €"] = f"{spendenbetrag:.2f}".replace(".", ",")
                neu["Steuer €"] = "0,00"
                neu["Brutto €"] = f"{spendenbetrag:.2f}".replace(".", ",")
                neu["Position(en) Anzahl"] = "1"
                neue_zeilen.append((neu, "spende"))
            elif artikel_0:
                neu = row.copy()
                neu["Position(en)"] = ", ".join(artikel_0)
                if len(mengen_0) == 1:
                    neu["Position(en) Anzahl"] = mengen_0[0]
                else:
                    neu["Position(en) Anzahl"] = "| ".join(mengen_0)
                neu["Position(en) USt %"] = "0"
                neu["Netto €"] = f"{brutto_0_sum:.2f}".replace(".", ",")
                neu["Steuer €"] = "0,00"
                neu["Brutto €"] = f"{brutto_0_sum:.2f}".replace(".", ",")
                neue_zeilen.append((neu, "spende"))

        output_df = pd.DataFrame([z[0] for z in neue_zeilen])
        output_path = args.output or (os.path.splitext(rechnung_path)[0] + "_sortiert.xlsx")
        output_df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        gruen = PatternFill("solid", start_color="C6EFCE")
        orange = PatternFill("solid", start_color="FFEB9C")
        blau = PatternFill("solid", start_color="DDEBF7")

        for i, (_, status) in enumerate(neue_zeilen, start=2):
            if status == "geteilt_ok":
                for c in ws[i]:
                    c.fill = gruen
            elif status == "geteilt_fehler":
                for c in ws[i]:
                    c.fill = orange
            elif status == "spende":
                for c in ws[i]:
                    c.fill = blau

        wb.save(output_path)
        
      # =======================
        # PHASE 2: AUSWERTUNG
        # =======================

        from collections import defaultdict
        from openpyxl.styles import Font

        # Excel erneut öffnen
        wb = load_workbook(output_path)
        ws = wb.worksheets[0]   # erstes Blatt = gesplittete Rechnungen

        # Header ermitteln
        header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

        def get(row, name):
            return row[header[name] - 1].value

        # Artikelliste vorbereiten
        artikel_lookup = {}
        for _, a in artikel_df.iterrows():
            artikel_lookup[normalize(a["Name*"])] = {
                "Kategorie": a["Kategorie"],
                "USt": "".join(c for c in str(a["USt.*"]) if c.isdigit()),
                "Netto": parse_decimal(a["Verkaufspreis Netto"]),
                "Brutto": parse_decimal(a["Verkaufspreis Brutto*"])
            }
        artikel_namen_phase2 = sorted(artikel_lookup.keys(), key=len, reverse=True)

        # Auswertung: Kategorie → Zahlungsart → Summen
        auswertung = defaultdict(lambda: defaultdict(lambda: {
            "Netto": Decimal("0.00"),
            "Steuer": Decimal("0.00"),
            "Brutto": Decimal("0.00")
        }))

        # Gesplittete Rechnungen durchlaufen
        for row in ws.iter_rows(min_row=2):
            zahlung = get(row, "Zahlungsart")
            ust_zeile = str(get(row, "Position(en) USt %")).strip()

            if ust_zeile == "?":
                continue

            pos_text = normalize(strip_quantities(get(row, "Position(en)")))
            mengen_text = str(get(row, "Position(en) Anzahl"))
            mengen = [int(x.strip()) for x in mengen_text.split("|") if x.strip()]

            # Artikel positionsgleich zerlegen
            artikel_seq, _ = extract_positions_sequential(pos_text, artikel_namen_phase2)
            if len(artikel_seq) < len(mengen):
                fallback_parts = [
                    normalize(strip_quantities(part))
                    for part in str(get(row, "Position(en)")).split(",")
                ]
                artikel_seq = [part for part in fallback_parts if part]
            if not mengen:
                mengen = [1] * len(artikel_seq)

            row_items = []

            for index, artikel_name in enumerate(artikel_seq):
                menge = mengen[index] if index < len(mengen) else 1
                info = artikel_lookup.get(artikel_name)
                if not info:
                    continue

                ust = info["USt"]

                if ust == "0":
                    kategorie = "Spende"
                    brutto = quantize_money(info["Brutto"] * Decimal(menge))
                    netto = brutto
                    steuer = Decimal("0.00")
                else:
                    kategorie = f"{info['Kategorie']} {ust}%"
                    steuersatz = Decimal(int(ust)) / Decimal("100")
                    brutto = quantize_money(info["Brutto"] * Decimal(menge))
                    netto = quantize_money(brutto / (Decimal("1") + steuersatz))
                    steuer = quantize_money(brutto - netto)

                row_items.append((kategorie, netto, steuer, brutto))

            if not row_items:
                continue

            row_netto = parse_decimal(get(row, "Netto €"))
            row_steuer = parse_decimal(get(row, "Steuer €"))
            row_brutto = parse_decimal(get(row, "Brutto €"))

            sum_netto = sum(item[1] for item in row_items)
            sum_steuer = sum(item[2] for item in row_items)
            sum_brutto = sum(item[3] for item in row_items)

            diff_netto = quantize_money(row_netto - sum_netto)
            diff_steuer = quantize_money(row_steuer - sum_steuer)
            diff_brutto = quantize_money(row_brutto - sum_brutto)

            kategorie, netto, steuer, brutto = row_items[-1]
            row_items[-1] = (
                kategorie,
                quantize_money(netto + diff_netto),
                quantize_money(steuer + diff_steuer),
                quantize_money(brutto + diff_brutto),
            )

            for kategorie, netto, steuer, brutto in row_items:
                auswertung[kategorie][zahlung]["Netto"] += netto
                auswertung[kategorie][zahlung]["Steuer"] += steuer
                auswertung[kategorie][zahlung]["Brutto"] += brutto

        # Altes Auswertungsblatt entfernen
        if "Auswertung" in wb.sheetnames:
            del wb["Auswertung"]

        ws_out = wb.create_sheet("Auswertung")

        # Header
        ws_out.append([
            "Kategorie",
            "Zahlungsart",
            "SUM von Netto",
            "SUM von Steuer",
            "SUM von Brutto"
        ])

        for c in ws_out[1]:
            c.font = Font(bold=True)

        # Daten schreiben
        for kategorie, zahlungen in auswertung.items():
            for zahlung, summen in zahlungen.items():
                ws_out.append([
                    kategorie,
                    zahlung,
                    float(quantize_money(summen["Netto"])),
                    float(quantize_money(summen["Steuer"])),
                    float(quantize_money(summen["Brutto"])),
                ])
                
        # =======================
        # GESAMTSUMME
        # =======================

        gesamt_netto = Decimal("0.00")
        gesamt_steuer = Decimal("0.00")
        gesamt_brutto = Decimal("0.00")

        for zahlungen in auswertung.values():
            for summen in zahlungen.values():
                gesamt_netto += summen["Netto"]
                gesamt_steuer += summen["Steuer"]
                gesamt_brutto += summen["Brutto"]

        ws_out.append([
            "GESAMT",
            "",
            float(quantize_money(gesamt_netto)),
            float(quantize_money(gesamt_steuer)),
            float(quantize_money(gesamt_brutto)),
        ])

        for cell in ws_out[ws_out.max_row]:
            cell.font = Font(bold=True)

        # Vergleich mit Blatt 1 Summen
        sheet_netto = Decimal("0.00")
        sheet_steuer = Decimal("0.00")
        sheet_brutto = Decimal("0.00")
        for row in ws.iter_rows(min_row=2):
            sheet_netto += parse_decimal(get(row, "Netto €"))
            sheet_steuer += parse_decimal(get(row, "Steuer €"))
            sheet_brutto += parse_decimal(get(row, "Brutto €"))

        ws_out.append([
            "GESAMT Blatt1",
            "",
            float(quantize_money(sheet_netto)),
            float(quantize_money(sheet_steuer)),
            float(quantize_money(sheet_brutto)),
        ])
        ws_out.append([
            "DIFF Blatt1",
            "",
            float(quantize_money(gesamt_netto - sheet_netto)),
            float(quantize_money(gesamt_steuer - sheet_steuer)),
            float(quantize_money(gesamt_brutto - sheet_brutto)),
        ])

        # Speichern
        wb.save(output_path)



        if not args.no_gui:
            from tkinter import messagebox

            antwort = messagebox.askyesno(
                "Fertig",
                "Die Datei wurde erfolgreich erstellt.\n\nMöchtest du die Datei jetzt öffnen?",
            )

            if antwort:
                if sys.platform.startswith("win"):
                    os.startfile(output_path)
                elif sys.platform.startswith("darwin"):
                    subprocess.run(["open", output_path])
                else:
                    subprocess.run(["xdg-open", output_path])

    except Exception as e:
        if not args.no_gui:
            from tkinter import messagebox

            messagebox.showerror("Fehler", str(e))
        print(traceback.format_exc())


if __name__ == "__main__":
    main()
