import pandas as pd
import subprocess
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import traceback
from decimal import Decimal, ROUND_HALF_EVEN, getcontext
getcontext().prec = 12

def bank_round(value, ndigits=2):
    q = Decimal("1." + "0" * ndigits)
    return float(Decimal(str(value)).quantize(q, rounding=ROUND_HALF_EVEN))

TOLERANZ = 0.01


def normalize(text):
    return " ".join(str(text).lower().strip().split())


def to_float(value):
    return float(str(value).replace(",", "."))


def extract_positions_sequential(text, artikel_namen):
    rest = normalize(text)
    gefunden = []

    while rest:
        match = None
        for name in artikel_namen:
            if rest.startswith(name):
                match = name
                break

        if not match:
            break

        gefunden.append(match)
        rest = rest[len(match):].lstrip()
        if rest.startswith(","):
            rest = rest[1:].lstrip()

    return gefunden, rest


def main():
    root = tk.Tk()
    root.withdraw()

    rechnung_path = filedialog.askopenfilename(
        title="Wähle die Rechnungsdatei aus",
        filetypes=[("CSV Dateien", "*.csv")]
    )
    if not rechnung_path:
        return

    artikel_path = filedialog.askopenfilename(
        title="Bitte Artikelliste auswählen",
        filetypes=[("CSV Dateien", "*.csv")]
    )
    if not artikel_path:
        return

    try:
        rechnung_df = pd.read_csv(rechnung_path, sep=";", dtype=str)
        artikel_df = pd.read_csv(artikel_path, sep=";", dtype=str)

        artikel_df["Name_norm"] = artikel_df["Name*"].apply(normalize)

        artikel_namen = sorted(
            artikel_df["Name_norm"].tolist(),
            key=len,
            reverse=True
        )

        artikel_map = {
            row["Name_norm"]: row
            for _, row in artikel_df.iterrows()
        }

        neue_zeilen = []

        for _, row in rechnung_df.iterrows():
            ust_werte = {
                u.strip()
                for u in str(row["Position(en) USt %"]).split(",")
                if u.strip()
            }

            if len(ust_werte) <= 1:
                neue_zeilen.append((row, None))
                continue

            artikel_seq, _ = extract_positions_sequential(
                row["Position(en)"],
                artikel_namen
            )

            mengen = [int(x.strip()) for x in row["Position(en) Anzahl"].split("|")]

            artikel_7 = []
            artikel_19 = []
            
            mengen_7 = []
            mengen_19 = []

            netto_7 = 0.0
            netto_19 = 0.0

            for artikel_norm, menge in zip(artikel_seq, mengen):
                daten = artikel_map.get(artikel_norm)
                if daten is None:
                    continue

                ust_num = "".join(c for c in str(daten["USt.*"]) if c.isdigit())

                if ust_num == "0":
                    continue  # Spende

                preis_netto = bank_round(
                    to_float(daten["Verkaufspreis Netto"]) * menge,
                    2
                )
                    
                original_name = daten["Name*"]

                if ust_num == "7":
                    artikel_7.append(f"{original_name} ({menge}x)")
                    mengen_7.append(str(menge))
                    netto_7 += preis_netto

                elif ust_num == "19":
                    artikel_19.append(f"{original_name} ({menge}x)")
                    mengen_19.append(str(menge))
                    netto_19 += preis_netto

            original_brutto = to_float(row["Brutto €"])

            berechnetes_brutto = round(
                (netto_7 + netto_7 * 0.07) +
                (netto_19 + netto_19 * 0.19),
                2
            )

            spendenbetrag = round(original_brutto - berechnetes_brutto, 2)

            brutto_ok = abs(original_brutto - berechnetes_brutto - spendenbetrag) <= TOLERANZ
            status = "geteilt_ok" if brutto_ok else "geteilt_fehler"

            if artikel_7:
                steuer_7 = bank_round(netto_7 * 0.07, 2)
                brutto_7 = bank_round(netto_7 + steuer_7, 2)

                neu = row.copy()
                neu["Position(en)"] = ", ".join(artikel_7)
                
                if len(mengen_7) == 1:
                    neu["Position(en) Anzahl"] = mengen_7[0]
                else:
                    neu["Position(en) Anzahl"] = "| ".join(mengen_7)
                
                neu["Position(en) USt %"] = "7"
                neu["Netto €"] = f"{netto_7:.2f}".replace(".", ",")
                neu["Steuer €"] = f"{steuer_7:.2f}".replace(".", ",")
                neu["Brutto €"] = f"{brutto_7:.2f}".replace(".", ",")
                neue_zeilen.append((neu, status))

            if artikel_19:
                steuer_19 = bank_round(netto_19 * 0.19, 2)
                brutto_19 = bank_round(netto_19 + steuer_19, 2)

                neu = row.copy()
                neu["Position(en)"] = ", ".join(artikel_19)
                
                if len(mengen_19) == 1:
                    neu["Position(en) Anzahl"] = mengen_19[0]
                else:
                    neu["Position(en) Anzahl"] = "| ".join(mengen_19)
                
                neu["Position(en) USt %"] = "19"
                neu["Netto €"] = f"{netto_19:.2f}".replace(".", ",")
                neu["Steuer €"] = f"{steuer_19:.2f}".replace(".", ",")
                neu["Brutto €"] = f"{brutto_19:.2f}".replace(".", ",")
                neue_zeilen.append((neu, status))

            if abs(spendenbetrag) > TOLERANZ:
                neu = row.copy()
                neu["Position(en)"] = "Spende"
                neu["Position(en) USt %"] = "0"
                neu["Netto €"] = f"{spendenbetrag:.2f}".replace(".", ",")
                neu["Steuer €"] = "0,00"
                neu["Brutto €"] = f"{spendenbetrag:.2f}".replace(".", ",")
                neu["Position(en) Anzahl"] = "1"
                neue_zeilen.append((neu, "spende"))

        output_df = pd.DataFrame([z[0] for z in neue_zeilen])
        output_path = os.path.splitext(rechnung_path)[0] + "_sortiert.xlsx"
        output_df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        gruen = PatternFill("solid", start_color="C6EFCE")
        orange = PatternFill("solid", start_color="FFEB9C")
        blau = PatternFill("solid", start_color="DDEBF7")

        for i, (_, status) in enumerate(neue_zeilen, start=2):
            if status == "geteilt_ok":
                for c in ws[i]:
                    c.fill = gruen
            elif status == "geteilt_fehler":
                for c in ws[i]:
                    c.fill = orange
            elif status == "spende":
                for c in ws[i]:
                    c.fill = blau

        wb.save(output_path)
        
      # =======================
        # PHASE 2: AUSWERTUNG
        # =======================

        from collections import defaultdict
        from openpyxl.styles import Font

        # Excel erneut öffnen
        wb = load_workbook(output_path)
        ws = wb.worksheets[0]   # erstes Blatt = gesplittete Rechnungen

        # Header ermitteln
        header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

        def get(row, name):
            return row[header[name] - 1].value

        # Artikelliste vorbereiten
        artikel_lookup = {}
        for _, a in artikel_df.iterrows():
            artikel_lookup[normalize(a["Name*"])] = {
                "Kategorie": a["Kategorie"],
                "USt": "".join(c for c in str(a["USt.*"]) if c.isdigit()),
                "Netto": to_float(a["Verkaufspreis Netto"])
            }

        # Auswertung: Kategorie → Zahlungsart → Summen
        auswertung = defaultdict(lambda: defaultdict(lambda: {
            "Netto": 0.0,
            "Steuer": 0.0,
            "Brutto": 0.0
        }))

        # Gesplittete Rechnungen durchlaufen
        for row in ws.iter_rows(min_row=2):
            zahlung = get(row, "Zahlungsart")
            ust_zeile = str(get(row, "Position(en) USt %")).strip()

            if ust_zeile == "?":
                continue

            pos_text = normalize(get(row, "Position(en)"))
            mengen = [int(x.strip()) for x in str(get(row, "Position(en) Anzahl")).split("|")]

            # Artikel positionsgleich zerlegen
            rest = pos_text
            artikel_seq = []

            while rest:
                match = None
                for name in artikel_lookup:
                    if rest.startswith(name):
                        match = name
                        break
                if not match:
                    break

                artikel_seq.append(match)
                rest = rest[len(match):].lstrip()
                if rest.startswith(","):
                    rest = rest[1:].lstrip()

            # Artikelweise auswerten
            for artikel_name, menge in zip(artikel_seq, mengen):
                info = artikel_lookup.get(artikel_name)
                if not info:
                    continue

                ust = info["USt"]

                if ust == "0":
                    kategorie = "Spende"
                    netto = to_float(get(row, "Netto €"))
                    steuer = 0.0
                    brutto = netto
                else:
                    kategorie = f"{info['Kategorie']} {ust}%"
                    netto = info["Netto"] * menge
                    steuer = round(netto * int(ust) / 100, 2)
                    brutto = round(netto + steuer, 2)

                auswertung[kategorie][zahlung]["Netto"] += netto
                auswertung[kategorie][zahlung]["Steuer"] += steuer
                auswertung[kategorie][zahlung]["Brutto"] += brutto

        # Altes Auswertungsblatt entfernen
        if "Auswertung" in wb.sheetnames:
            del wb["Auswertung"]

        ws_out = wb.create_sheet("Auswertung")

        # Header
        ws_out.append([
            "Kategorie",
            "Zahlungsart",
            "SUM von Netto",
            "SUM von Steuer",
            "SUM von Brutto"
        ])

        for c in ws_out[1]:
            c.font = Font(bold=True)

        # Daten schreiben
        for kategorie, zahlungen in auswertung.items():
            for zahlung, summen in zahlungen.items():
                ws_out.append([
                    kategorie,
                    zahlung,
                    round(summen["Netto"], 2),
                    round(summen["Steuer"], 2),
                    round(summen["Brutto"], 2),
                ])
                
        # =======================
        # GESAMTSUMME
        # =======================

        gesamt_netto = 0.0
        gesamt_steuer = 0.0
        gesamt_brutto = 0.0

        for zahlungen in auswertung.values():
            for summen in zahlungen.values():
                gesamt_netto += summen["Netto"]
                gesamt_steuer += summen["Steuer"]
                gesamt_brutto += summen["Brutto"]

        ws_out.append([
            "GESAMT",
            "",
            round(gesamt_netto, 2),
            round(gesamt_steuer, 2),
            round(gesamt_brutto, 2),
        ])

        # optional: fett markieren
        for cell in ws_out[ws_out.max_row]:
            cell.font = Font(bold=True)

        # Speichern
        wb.save(output_path)



        antwort = messagebox.askyesno(
            "Fertig",
            "Die Datei wurde erfolgreich erstellt.\n\nMöchtest du die Datei jetzt öffnen?"
        )

        if antwort:
            if sys.platform.startswith("win"):
                os.startfile(output_path)
            elif sys.platform.startswith("darwin"):
                subprocess.run(["open", output_path])
            else:
                subprocess.run(["xdg-open", output_path])

    except Exception as e:
        messagebox.showerror("Fehler", str(e))
        print(traceback.format_exc())


if __name__ == "__main__":
    main()
